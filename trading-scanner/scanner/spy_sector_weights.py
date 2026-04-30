"""SPY sector cap-weights ingest — bug #1087 Phase 1.

Pulls the daily State Street SPDR SPY holdings file, aggregates weight
by GICS sector, and upserts one row per (date, sector) into
public.spy_sector_weights.

Pipeline name:  market-spy_sector_weights-daily
Cadence:        weekday daily, 06:00 ET
Owner:          Data Steward

Source URL:
    https://www.ssga.com/us/en/intermediary/library-content/products/fund-data/etfs/us/holdings-daily-us-en-spy.xlsx

The file is published by SSGA each market day and is free + license-tier
"public". Layout: a few metadata rows on top, a header row containing the
literal "Sector" column, then one row per holding. Cash & derivatives rows
have no sector and are skipped from the aggregation.

Failure modes:
  - Non-200 HTTP / corrupted file: raise so the GH workflow shows red and
    DAILY-HOME-SMOKE downstream can flag staleness.
  - Two consecutive failures filed as P1 bug per pipeline_schedule.yml
    (see Data Steward kickoff doc).

Acceptance gate for #1087 Phase 1:
  - Row count > 0 in public.spy_sector_weights for the run date
  - Sector weights sum to ~1.0 (within 0.01 tolerance)
"""

from __future__ import annotations

import io
import logging
import os
from datetime import date as date_cls, datetime, timezone
from typing import Any

import requests

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

# SSGA publishes the SPY holdings file at the URL below. URL has been stable
# since 2018; if SSGA changes the path this script will 404 and the GH
# workflow shows red — file a bug at that point and update the URL.
SSGA_SPY_HOLDINGS_URL = (
    "https://www.ssga.com/us/en/intermediary/library-content/products/"
    "fund-data/etfs/us/holdings-daily-us-en-spy.xlsx"
)

REQUEST_TIMEOUT = 30
MAX_RETRIES = 3
RETRY_BACKOFF = (1.0, 2.0, 4.0)

# Tolerance on the sector-sum sanity check. SSGA holdings include a small
# residual cash / unsettled-trade weight (~0.01% on most days), so allow
# 0.99 ≤ sum ≤ 1.01.
SECTOR_SUM_TOLERANCE = 0.01


# ---------------------------------------------------------------------------
# HTTP fetch
# ---------------------------------------------------------------------------

def _fetch_holdings_file() -> bytes:
    """Download the SSGA SPY holdings xlsx with retry + backoff."""
    last_err: Exception | None = None
    for attempt, backoff in enumerate(RETRY_BACKOFF[:MAX_RETRIES], start=1):
        try:
            logger.info("Fetching SSGA SPY holdings (attempt %d/%d)", attempt, MAX_RETRIES)
            resp = requests.get(
                SSGA_SPY_HOLDINGS_URL,
                timeout=REQUEST_TIMEOUT,
                headers={"User-Agent": "macrotilt-data-steward/1.0"},
            )
            resp.raise_for_status()
            if not resp.content or len(resp.content) < 1000:
                raise ValueError(
                    f"SSGA file too small ({len(resp.content)} bytes) — likely an error page"
                )
            return resp.content
        except Exception as e:  # noqa: BLE001
            last_err = e
            logger.warning("SSGA fetch failed: %s — sleeping %ss", e, backoff)
            if attempt < MAX_RETRIES:
                import time
                time.sleep(backoff)
    assert last_err is not None
    raise last_err


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

def _parse_holdings_xlsx(content: bytes) -> tuple[date_cls, dict[str, float]]:
    """Parse SSGA's SPY holdings xlsx → (as_of_date, {sector: weight_pct}).

    Aggregation: sums Weight column by Sector. Skips rows with empty Sector
    (cash / unsettled / fund-of-fund placeholders). Returns weights as
    decimals in [0, 1]; SSGA publishes them as percentages (0–100) so we
    divide.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "openpyxl not installed — add `openpyxl>=3.1` to requirements.txt"
        ) from e

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # Find the as-of date in the metadata block and the header row.
    as_of: date_cls | None = None
    header_row_idx: int | None = None
    sector_col: int | None = None
    weight_col: int | None = None

    rows = list(ws.iter_rows(values_only=True))
    for idx, row in enumerate(rows):
        cells = [str(c) if c is not None else "" for c in row]
        joined_lc = " ".join(cells).lower()
        # Metadata "as of" line
        if as_of is None and "as of" in joined_lc:
            for cell in row:
                if isinstance(cell, datetime):
                    as_of = cell.date()
                    break
                if isinstance(cell, date_cls):
                    as_of = cell
                    break
                if isinstance(cell, str):
                    # Try to parse common SSGA formats
                    for fmt in ("%b-%d-%Y", "%B %d, %Y", "%m/%d/%Y", "%Y-%m-%d"):
                        try:
                            as_of = datetime.strptime(cell.strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                    if as_of is not None:
                        break
        # Header row
        lower_cells = [c.lower() for c in cells]
        if "sector" in lower_cells and "weight" in " ".join(lower_cells):
            # Many SSGA files label the col "Weight (%)"; match prefix.
            for ci, c in enumerate(lower_cells):
                if c == "sector":
                    sector_col = ci
                if c.startswith("weight"):
                    weight_col = ci
            if sector_col is not None and weight_col is not None:
                header_row_idx = idx
                break

    if header_row_idx is None or sector_col is None or weight_col is None:
        raise RuntimeError("Could not locate Sector / Weight columns in SSGA xlsx")

    # Default as_of to today if SSGA didn't print a parseable as-of date.
    if as_of is None:
        as_of = datetime.now(timezone.utc).date()
        logger.warning("SSGA file did not contain a parseable 'as of' date — using today (UTC)")

    # Aggregate weights by sector
    by_sector: dict[str, float] = {}
    for row in rows[header_row_idx + 1 :]:
        if row is None:
            continue
        sector_val = row[sector_col]
        weight_val = row[weight_col]
        if sector_val is None or weight_val is None:
            continue
        sector = str(sector_val).strip()
        if not sector or sector.lower() in ("unassigned", "-", "n/a"):
            continue
        try:
            weight = float(weight_val)
        except (TypeError, ValueError):
            continue
        # SSGA publishes Weight as percent (e.g. 27.3 means 27.3%).
        # If we ever see a value > 1.5 we know it's percent — divide by 100.
        # If <= 1.0 across the board it's already decimal — leave as-is.
        # Decide once on the first numeric we see.
        by_sector[sector] = by_sector.get(sector, 0.0) + weight

    # Decide unit: total > 1.5 means percent → divide
    total = sum(by_sector.values())
    if total > 1.5:
        by_sector = {k: v / 100.0 for k, v in by_sector.items()}
        total = total / 100.0

    if not by_sector:
        raise RuntimeError("No sector rows parsed from SSGA xlsx")
    if abs(total - 1.0) > SECTOR_SUM_TOLERANCE:
        raise RuntimeError(
            f"Sector weights sum to {total:.4f}, outside tolerance "
            f"{1 - SECTOR_SUM_TOLERANCE:.2f}–{1 + SECTOR_SUM_TOLERANCE:.2f}"
        )

    return as_of, by_sector


# ---------------------------------------------------------------------------
# Supabase upsert
# ---------------------------------------------------------------------------

def _get_supabase_client():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError(
            "SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY not set — cannot write"
        )
    from supabase import create_client  # type: ignore

    return create_client(url, key)


def _upsert_rows(client, as_of: date_cls, by_sector: dict[str, float]) -> int:
    rows = [
        {
            "date": as_of.isoformat(),
            "sector": sector,
            "weight_pct": round(weight, 5),
            "source": "ssga_spdr_spy_holdings",
        }
        for sector, weight in by_sector.items()
    ]
    resp = (
        client.table("spy_sector_weights")
        .upsert(rows, on_conflict="date,sector")
        .execute()
    )
    if getattr(resp, "data", None) is None:
        raise RuntimeError(f"upsert returned no data: {resp}")
    return len(resp.data)


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

def run_spy_sector_weights(*, dry_run: bool = False) -> dict[str, Any]:
    """Fetch + parse + upsert. Returns a JSON-serializable summary."""
    started_at = datetime.now(timezone.utc)
    content = _fetch_holdings_file()
    as_of, by_sector = _parse_holdings_xlsx(content)
    sector_total = sum(by_sector.values())

    summary: dict[str, Any] = {
        "pipeline": "market-spy_sector_weights-daily",
        "as_of": as_of.isoformat(),
        "sectors": len(by_sector),
        "weight_sum": round(sector_total, 5),
        "dry_run": dry_run,
        "source": SSGA_SPY_HOLDINGS_URL,
        "started_at": started_at.isoformat(),
    }

    if dry_run:
        summary["sector_breakdown"] = {k: round(v, 5) for k, v in by_sector.items()}
        summary["rows_upserted"] = 0
    else:
        client = _get_supabase_client()
        summary["rows_upserted"] = _upsert_rows(client, as_of, by_sector)

    summary["duration_seconds"] = round(
        (datetime.now(timezone.utc) - started_at).total_seconds(), 2
    )
    return summary
