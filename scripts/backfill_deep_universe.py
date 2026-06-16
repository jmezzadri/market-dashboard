"""backfill_deep_universe.py — deep (~1996+) daily price history for the Asset
Tilt backtest universe, into public.prices_eod from yfinance (free deep source,
the LESSONS 7.2 bootstrap pattern). One-shot / re-runnable; idempotent upsert on
(ticker, trade_date). Massive owns the recent forward window; this fills the deep
history the breadth recompute + the 1996+ sector/engine backtest need.

Universe: 11 GICS sector ETFs + industry-group ETFs + defensive/benchmark ETFs +
the live S&P 500 (SPY) and Nasdaq-100 (QQQ) members (today's membership applied
back — matches the live breadth producer).
"""
import json, os, sys, time, io, re, csv
from datetime import date
from urllib.request import Request, urlopen
from urllib.parse import urlencode
from urllib.error import HTTPError

START_DATE = "1996-01-01"
END_DATE   = date.today().isoformat()
UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"

SECTOR_ETFS = ["XLK","XLF","XLE","XLI","XLY","XLP","XLV","XLU","XLB","XLRE","XLC"]
INDUSTRY_ETFS = ["SMH","SOXX","XBI","IBB","KRE","KBE","ITB","XHB","XOP","OIH","XME",
                 "XRT","JETS","IGV","CIBR","GDX","KIE","IYT","PAVE","XAR","TAN","XTL"]
ENGINE_ETFS = ["SPY","QQQ","TLT","IEF","SHY","BIL","GLD","LQD","HYG","AGG","IYR","MGK"]

def env(n):
    v=os.environ.get(n,"").strip()
    if not v: raise SystemExit(f"missing env var: {n}")
    return v

def spy_members():
    import requests
    from openpyxl import load_workbook
    r=requests.get("https://www.ssga.com/us/en/intermediary/library-content/products/fund-data/etfs/us/holdings-daily-us-en-spy.xlsx",headers={"User-Agent":UA},timeout=60)
    wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True); ws=wb.active
    rows=[list(row) for row in ws.iter_rows(values_only=True)]
    hi=next(i for i,row in enumerate(rows[:12]) if row and any(str(c).strip().lower()=='ticker' for c in row if c))
    ti=[str(c).strip() if c else '' for c in rows[hi]].index('Ticker')
    return sorted({row[ti].strip() for row in rows[hi+1:] if row and ti<len(row) and isinstance(row[ti],str) and row[ti].strip() and row[ti].strip()!='-'})

def qqq_members():
    import requests, pandas as pd
    try:
        r=requests.get("https://www.invesco.com/us/financial-products/etfs/holdings/main/holdings/0?audienceType=Investor&action=download&ticker=QQQ",headers={"User-Agent":UA},timeout=60)
        rowsc=list(csv.reader(io.StringIO(r.content.decode('utf-8-sig',errors='ignore'))))
        hi=next((i for i,row in enumerate(rowsc[:40]) if any('ticker' in str(c).lower() for c in row)),None)
        if hi is not None:
            ci=next((i for i,c in enumerate([str(c).strip().lower() for c in rowsc[hi]]) if 'ticker' in c),None)
            if ci is not None:
                syms=sorted({row[ci].strip() for row in rowsc[hi+1:] if len(row)>ci and row[ci].strip() and row[ci].strip()!='-'})
                if len(syms)>=90: return syms
    except Exception as e: print("NDX Invesco failed:",repr(e),"— Wikipedia fallback")
    r=requests.get("https://en.wikipedia.org/wiki/Nasdaq-100",headers={"User-Agent":UA},timeout=30); r.raise_for_status()
    for t in pd.read_html(io.StringIO(r.text)):
        for col in ("Ticker","Symbol"):
            if col in t.columns:
                syms=sorted({re.sub(r'[^A-Z.\-]','',str(x).upper()) for x in t[col].astype(str)})
                syms=[x for x in syms if 1<=len(x)<=6]
                if len(syms)>=90: return syms
    raise RuntimeError("NDX membership unavailable")

def _f(v):
    try:
        import math; f=float(v); return None if math.isnan(f) else round(f,6)
    except Exception: return None
def _i(v):
    try: return int(v)
    except Exception: return None

def supabase_upsert(url,key,rows,chunk=500):
    if not rows: return 0
    endpoint=f"{url.rstrip('/')}/rest/v1/prices_eod?{urlencode({'on_conflict':'ticker,trade_date'})}"
    hdr={"apikey":key,"Authorization":f"Bearer {key}","Content-Type":"application/json","Prefer":"resolution=merge-duplicates,return=minimal"}
    total=0
    for i in range(0,len(rows),chunk):
        sl=rows[i:i+chunk]; req=Request(endpoint,data=json.dumps(sl).encode(),method="POST")
        for h,v in hdr.items(): req.add_header(h,v)
        try:
            with urlopen(req,timeout=120) as r:
                if r.status>=300: raise RuntimeError(f"HTTP {r.status}")
                total+=len(sl)
        except HTTPError as e:
            raise RuntimeError(f"upsert HTTP {e.code}: {e.read().decode()[:200]}")
    return total

def fetch_yf(ticker):
    import yfinance as yf, pandas as pd
    yf_sym=ticker.replace(".","-")  # class shares: BRK.B/BF.B -> BRK-B/BF-B for yfinance
    df=yf.download(yf_sym,start=START_DATE,end=END_DATE,progress=False,auto_adjust=False,threads=False)
    if df is None or df.empty: return []
    if isinstance(df.columns,pd.MultiIndex): df.columns=df.columns.get_level_values(0)
    out=[]
    for d,row in df.iterrows():
        try: iso=d.strftime("%Y-%m-%d")
        except Exception: iso=str(d)[:10]
        c=_f(row.get("Close"))
        if c is None: continue
        out.append({"ticker":ticker,"trade_date":iso,"open":_f(row.get("Open")),"high":_f(row.get("High")),
                    "low":_f(row.get("Low")),"close":c,"volume":_i(row.get("Volume")),
                    "vwap":None,"transactions":None,"source":"yfinance-deep-backfill"})
    return out

def main():
    URL=env("SUPABASE_URL"); SK=env("SUPABASE_SERVICE_ROLE_KEY")
    try: members=sorted(set(spy_members())|set(qqq_members()))
    except Exception as e:
        print("MEMBER FETCH FAILED:",repr(e)); members=[]
    universe=sorted(set(SECTOR_ETFS+INDUSTRY_ETFS+ENGINE_ETFS+members))
    only=os.environ.get("ONLY","").strip()
    if only:
        universe=[t.strip() for t in only.split(",") if t.strip()]
        print("ONLY override — restricting universe to:",universe)
    print(f"Deep backfill {START_DATE}→{END_DATE}  |  {len(universe)} tickers "
          f"({len(members)} members + {len(set(SECTOR_ETFS+INDUSTRY_ETFS+ENGINE_ETFS))} ETFs)")
    grand=0; fails=[]; deep=0
    for i,t in enumerate(universe,1):
        try:
            rows=fetch_yf(t)
            if not rows: fails.append((t,"no data")); print(f"[{i}/{len(universe)}] {t}: no data"); continue
            n=supabase_upsert(URL,SK,rows); grand+=n
            first,last=rows[0]["trade_date"],rows[-1]["trade_date"]
            if first<="2006-12-31": deep+=1
            if i%25==0 or first<="2006-12-31": print(f"[{i}/{len(universe)}] {t}: {n} bars {first}→{last}")
        except Exception as e:
            fails.append((t,repr(e)[:80])); print(f"[{i}/{len(universe)}] {t}: FAIL {repr(e)[:80]}")
        time.sleep(0.4)  # gentle on Yahoo
    print(f"\nDONE. {grand} bars upserted across {len(universe)-len(fails)} tickers; "
          f"{deep} tickers reach 2006 or earlier. Failures: {len(fails)}")
    for t,why in fails[:40]: print("  fail",t,why)

if __name__=="__main__": main()
